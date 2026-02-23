#!/usr/bin/env python3
"""
miafpe_to_jsonld.py
====================
Converts a MIAFPE xlsx metadata template + its OWL ontology file into a
JSON-LD knowledge graph.

Usage
-----
    python miafpe_to_jsonld.py <template.xlsx> <ontology.owl> [output.jsonld]

If the output path is omitted the result is written to stdout.

Dependencies
------------
    pip install openpyxl rdflib
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, date

import openpyxl
from rdflib import Graph, Namespace, RDF, RDFS, OWL, XSD
from rdflib.term import URIRef, Literal, BNode


# ─── Namespaces ──────────────────────────────────────────────────────────────

MIAFPE_NS = "https://github.com/gryvity/MIAFPE/tree/master/ontology/miappexsosa.owl#"
MIAFPE    = Namespace(MIAFPE_NS)


# ─── OWL helpers ─────────────────────────────────────────────────────────────

def _cardinality_of(g: Graph, restriction: BNode) -> dict:
    """
    Return a dict with keys 'min', 'max', 'exact' (None if not present)
    and 'property' (URIRef) for an owl:Restriction blank node.
    """
    prop  = g.value(restriction, OWL.onProperty)
    exact = g.value(restriction, OWL.qualifiedCardinality)   # owl:qualifiedCardinality
    exact = exact or g.value(restriction, OWL.cardinality)   # owl:cardinality
    min_c = g.value(restriction, OWL.minQualifiedCardinality)
    min_c = min_c or g.value(restriction, OWL.minCardinality)
    max_c = g.value(restriction, OWL.maxQualifiedCardinality)
    max_c = max_c or g.value(restriction, OWL.maxCardinality)
    return {
        "property": prop,
        "exact":    int(str(exact)) if exact is not None else None,
        "min":      int(str(min_c)) if min_c is not None else None,
        "max":      int(str(max_c)) if max_c is not None else None,
    }


def parse_owl_class(g: Graph, class_uri: URIRef) -> dict:
    """
    Parse a single owl:Class and return a summary dict.

    Returns
    -------
    {
        "uri":         str,
        "properties":  {local_name: {"uri": str, "multi": bool, "label": str}},
    }
    """
    # Collect all rdfs:subClassOf targets
    restrictions = {}   # prop_uri -> {"multi": bool, "label": str}

    for sc in g.objects(class_uri, RDFS.subClassOf):
        # Only care about blank-node restrictions
        if not isinstance(sc, BNode):
            continue
        if (sc, RDF.type, OWL.Restriction) not in g:
            continue

        card = _cardinality_of(g, sc)
        prop = card["property"]
        if prop is None:
            continue

        prop_str = str(prop)

        # Decide whether multiple values are allowed:
        # • exact == 1  → single value
        # • max   == 1  → single value
        # • otherwise   → list (multi=True)
        if card["exact"] == 1 or card["max"] == 1:
            multi = False
        else:
            multi = True

        # Try to find a human label from owl:Axiom annotations
        label = None
        for axiom in g.subjects(OWL.annotatedSource, class_uri):
            if (axiom, OWL.annotatedTarget, sc) in g:
                lbl = g.value(axiom, RDFS.label)
                if lbl:
                    label = str(lbl)

        # Local name is the fragment after '#'
        local = prop_str.split("#")[-1]

        if prop_str not in restrictions:
            restrictions[prop_str] = {"local": local, "multi": multi, "label": label}
        else:
            # If any restriction says multi, honour that
            if multi:
                restrictions[prop_str]["multi"] = True

    return {
        "uri":        str(class_uri),
        "properties": restrictions,
    }


def load_owl(owl_path: str) -> dict:
    """
    Parse the OWL file and return a dict keyed by class local-name.
    {
        "Investigation": { "uri": ..., "properties": {...} },
        ...
    }
    """
    g = Graph()
    g.parse(owl_path)

    classes = {}
    for cls in g.subjects(RDF.type, OWL.Class):
        if not isinstance(cls, URIRef):
            continue
        local = str(cls).split("#")[-1]
        classes[local] = parse_owl_class(g, cls)

    return classes


# ─── XLSX helpers ─────────────────────────────────────────────────────────────

def _coerce_value(v):
    """Convert openpyxl cell values to JSON-friendly Python types."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float):
        # Return int when the float is whole (e.g. 3.0 → 3)
        return int(v) if v == int(v) else v
    return str(v)


def load_xlsx(xlsx_path: str):
    """
    Returns
    -------
    info_rows : list of dicts  –  one dict per row in the 'Information' sheet
                                   {"sheet", "property", "ontology_name",
                                    "miafpe_iri", "cardinality", "range"}
    data_sheets: {sheet_name: {"headers": [...], "rows": [[...], ...]}}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── Information sheet ────────────────────────────────────────────────────
    info_ws = wb["Information"]
    info_header = [c.value for c in next(info_ws.iter_rows(min_row=1, max_row=1))]

    col_idx = {h: i for i, h in enumerate(info_header) if h}
    # Expected columns (case-insensitive search for robustness)
    col_map = {}
    for key, candidates in {
        "sheet":        ["Sheet Name"],
        "property":     ["Property"],
        "ontology_name":["Ontology Name"],
        "cardinality":  ["Cardinality"],
        "range":        ["Range"],
    }.items():
        for cand in candidates:
            if cand in col_idx:
                col_map[key] = col_idx[cand]
                break

    info_rows = []
    current_sheet = None
    for row in info_ws.iter_rows(min_row=2, values_only=True):
        sheet_val = row[col_map["sheet"]] if "sheet" in col_map else None
        if sheet_val:
            current_sheet = sheet_val

        prop   = row[col_map["property"]]     if "property"      in col_map else None
        ont    = row[col_map["ontology_name"]] if "ontology_name" in col_map else None
        card   = row[col_map["cardinality"]]   if "cardinality"   in col_map else None
        rng    = row[col_map["range"]]         if "range"         in col_map else None

        if prop and ont:
            info_rows.append({
                "sheet":         current_sheet,
                "property":      prop,
                "ontology_name": ont,   # e.g. "hasIdentifier"
                "cardinality":   card,
                "range":         rng,
            })

    # ── Data sheets ──────────────────────────────────────────────────────────
    data_sheets = {}
    for name in wb.sheetnames:
        if name == "Information":
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_coerce_value(c) for c in rows[0]]
        data_rows = []
        for row in rows[1:]:
            coerced = [_coerce_value(c) for c in row]
            # Skip fully-empty rows
            if any(v is not None for v in coerced):
                data_rows.append(coerced)
        data_sheets[name] = {"headers": headers, "rows": data_rows}

    return info_rows, data_sheets


# ─── JSON-LD builder ─────────────────────────────────────────────────────────

# Cardinality strings from the Information sheet that mean "multi"
_MULTI_CARDS = {"Some", None, ""}


def _is_multi_from_info(card_str: str) -> bool:
    """
    Return True when the cardinality from the Information sheet allows
    multiple values.
      Exactly(1), Max(1)  → single
      anything else (including None) → multi / list
    """
    if card_str is None:
        return True
    c = card_str.strip()
    return c not in ("Exactly(1)", "Max(1)")


def build_jsonld(info_rows: list, data_sheets: dict, owl_classes: dict,
                 base_ns: str = MIAFPE_NS) -> dict:
    """
    Build the full JSON-LD document.
    """

    # Build a lookup: sheet → {column_header → info_row}
    sheet_col_map: dict[str, dict[str, dict]] = {}
    for row in info_rows:
        sheet = row["sheet"]
        if sheet not in sheet_col_map:
            sheet_col_map[sheet] = {}
        sheet_col_map[sheet][row["property"]] = row

    # ── Context ──────────────────────────────────────────────────────────────
    context = {
        "@vocab": base_ns,
        "xsd":    "http://www.w3.org/2001/XMLSchema#",
    }

    # ── Nodes ────────────────────────────────────────────────────────────────
    graph = []

    for sheet_name, sheet_data in data_sheets.items():
        headers  = sheet_data["headers"]
        rows     = sheet_data["rows"]

        # Map headers to info rows for this sheet
        col_info = sheet_col_map.get(sheet_name, {})

        # Find the column that holds the unique identifier (ontology_name == "hasIdentifier")
        id_col_idx = None
        for i, h in enumerate(headers):
            meta = col_info.get(h)
            if meta and meta["ontology_name"] == "hasIdentifier":
                id_col_idx = i
                break

        # Lookup OWL class for multi/single decisions (optional – fallback to info sheet)
        owl_cls = owl_classes.get(sheet_name, {})
        owl_props = owl_cls.get("properties", {})

        for row in rows:
            node = {
                "@type": base_ns + sheet_name,
            }

            # Set @id from the identifier column
            if id_col_idx is not None:
                raw_id = row[id_col_idx]
                if raw_id:
                    node["@id"] = str(raw_id)

            for i, header in enumerate(headers):
                if header is None:
                    continue
                meta  = col_info.get(header)
                value = row[i] if i < len(row) else None

                if value is None:
                    continue

                # Determine the property name and IRI
                ont_name = meta["ontology_name"] if meta else header
                prop_iri = base_ns + ont_name

                # Determine multiplicity:
                # 1. OWL restriction (most precise)
                # 2. Information sheet cardinality string
                # 3. Default to single
                owl_prop_info = owl_props.get(prop_iri, {})
                if owl_prop_info:
                    multi = owl_prop_info.get("multi", False)
                elif meta:
                    multi = _is_multi_from_info(meta.get("cardinality"))
                else:
                    multi = False

                # Values can be semicolon-separated lists in the xlsx
                raw_val = str(value)
                if ";" in raw_val and multi:
                    parts = [p.strip() for p in raw_val.split(";") if p.strip()]
                else:
                    parts = [raw_val]

                # Try to detect date/numeric types from range hint
                rng = (meta or {}).get("range", "") or ""
                typed_parts = []
                for p in parts:
                    typed_parts.append(_type_value(p, rng))

                node[ont_name] = typed_parts if (multi or len(typed_parts) > 1) else typed_parts[0]

            graph.append(node)

    return {
        "@context": context,
        "@graph":   graph,
    }


def _type_value(raw: str, range_hint: str):
    """Add basic XSD typing based on the range hint from the Information sheet."""
    rh = (range_hint or "").lower()
    if "date" in rh or "time" in rh:
        return {"@type": "xsd:dateTime", "@value": raw}
    if "uri" in rh or raw.startswith("http"):
        return {"@id": raw}
    return raw


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert a MIAFPE xlsx template + OWL ontology to JSON-LD."
    )
    parser.add_argument("xlsx",  help="Path to the MIAFPE xlsx template")
    parser.add_argument("owl",   help="Path to the OWL ontology file")
    parser.add_argument("output", nargs="?", default=None,
                        help="Output JSON-LD file (default: stdout)")
    args = parser.parse_args()

    print(f"[1/3] Parsing OWL ontology: {args.owl}", file=sys.stderr)
    owl_classes = load_owl(args.owl)
    print(f"      Found {len(owl_classes)} classes: {list(owl_classes.keys())}",
          file=sys.stderr)

    print(f"[2/3] Reading xlsx template: {args.xlsx}", file=sys.stderr)
    info_rows, data_sheets = load_xlsx(args.xlsx)
    print(f"      Sheets with data: {[k for k, v in data_sheets.items() if v['rows']]}",
          file=sys.stderr)

    print("[3/3] Building JSON-LD …", file=sys.stderr)
    jsonld = build_jsonld(info_rows, data_sheets, owl_classes)

    out_str = json.dumps(jsonld, indent=2, ensure_ascii=False)

    if args.output:
        Path(args.output).write_text(out_str, encoding="utf-8")
        print(f"      Written to: {args.output}", file=sys.stderr)
    else:
        print(out_str)


if __name__ == "__main__":
    main()
